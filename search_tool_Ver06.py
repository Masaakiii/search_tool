import os
import pandas as pd
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter

selected_file_global = None

# 忽略 openpyxl 的警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def browse_folder():
    """讓使用者選擇資料夾"""
    folder = filedialog.askdirectory()
    if folder:
        folder_path_var.set(folder)

def display_results(file_name, sheet_name, results):
    """顯示選定檔案和工作表的搜尋結果"""
    # 清空表格
    for row in tree.get_children():
        tree.delete(row)
    
    # 檢查結果是否存在
    if file_name in results and sheet_name in results[file_name]:
        print(f"顯示結果: 檔案={file_name}, 工作表={sheet_name}")  # 調試訊息
        for row_data in results[file_name][sheet_name]:
            tree.insert("", "end", values=row_data)  # 插入每行資料
    else:
        print(f"無搜尋結果: 檔案={file_name}, 工作表={sheet_name}")  # 調試訊息
        messagebox.showinfo("提示", "無搜尋結果")

def export_results():
    if not grouped_results:
        messagebox.showerror("錯誤", "目前沒有搜尋結果可匯出！")
        return

    export_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="儲存搜尋結果"
    )

    if not export_path:
        return

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        for file_name, sheets in grouped_results.items():
            for sheet_name, rows in sheets.items():
                df = pd.DataFrame(rows, columns=[
                    "箱號", "JAN", "數量", "品名", "客編", "客戶",
                    "淨重", "毛重", "材積", "直派單號", "備註", "宅配單號", "宅配", "TIMES"
                ])
                sheet_title = f"{file_name}_{sheet_name}"[:31]
                df.to_excel(writer, sheet_name=sheet_title, index=False)

        # 💡 取得 openpyxl 的 workbook，調整每個工作表的欄寬
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 10)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    messagebox.showinfo("成功", f"結果已成功匯出到：\n{export_path}")

def search_keywords():
    """執行搜尋"""
    folder_path = folder_path_var.get()
    keyword1 = keyword1_var.get()
    keyword2 = keyword2_var.get()

    if not folder_path or not keyword1:
        messagebox.showerror("錯誤", "請選擇資料夾並輸入第一個搜尋條件！")
        return

    global grouped_results
    grouped_results = {}
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
            file_path = os.path.join(folder_path, file_name)
            try:
                # 讀取 Excel 檔案
                excel_data = pd.ExcelFile(file_path)
                
                # 遍歷每個工作表，僅搜尋名稱包含 "packing" 的工作表
                for sheet_name in excel_data.sheet_names:
                    if "packing" in sheet_name.lower():  # 忽略大小寫
                        df = excel_data.parse(sheet_name)
                        df = df.astype(str)  # 將所有值轉換為字串
                        
                        # 搜尋第一個關鍵字
                        rows_with_keyword1 = df.isin([keyword1]).any(axis=1)
                        if rows_with_keyword1.any():
                            for index, row in df[rows_with_keyword1].iterrows():
                                # 如果 keyword2 為空，直接輸出結果；否則檢查該行是否包含第二個關鍵字
                                if not keyword2 or keyword2 in row.values:
                                    result = [
                                        row.get("箱號", ""),  # 箱號
                                        row.get("JAN", ""),  # JAN
                                        row.get("數量", ""),  # 數量
                                        row.get("品名", ""),  # 品名
                                        row.get("客編", ""),  # 客編
                                        row.get("客戶", ""),  # 客戶
                                        row.get("淨重", ""),  # 淨重
                                        row.get("毛重", ""),  # 毛重
                                        row.get("材積", ""),  # 材積
                                        row.get("直派單號", ""),  # 直派單號
                                        row.get("備註", ""),  # 備註
                                        row.get("宅配單號", ""),  # 宅配單號
                                        row.get("宅配", ""),  # 宅配
                                        row.get("TIMES", ""),  # TIMES
                                    ]
                                    
                                    # 分組結果
                                    if file_name not in grouped_results:
                                        grouped_results[file_name] = {}
                                    if sheet_name not in grouped_results[file_name]:
                                        grouped_results[file_name][sheet_name] = []
                                    grouped_results[file_name][sheet_name].append(result)
            except Exception as e:
                grouped_results[file_name] = {"錯誤": [f"無法讀取檔案: {e}"]}

    # 顯示檔案名稱到左側
    file_listbox.delete(0, tk.END)
    for file_name in grouped_results.keys():
        file_listbox.insert(tk.END, file_name)
    
    # 綁定點擊事件，顯示中間的工作表名稱
    def on_file_select(event):
        """當選擇檔案時，顯示中間的工作表名稱"""
        global selected_file_global
        selected_index = file_listbox.curselection()
        if selected_index:
            selected_file_global = file_listbox.get(selected_index)  # 儲存選擇的檔案名稱
            print(f"選擇的檔案: {selected_file_global}")  # 調試訊息
            sheet_listbox.delete(0, tk.END)
            if selected_file_global in grouped_results:
                for sheet_name in grouped_results[selected_file_global].keys():
                    sheet_listbox.insert(tk.END, sheet_name)
            else:
                print(f"檔案 {selected_file_global} 無對應的工作表")  # 調試訊息
        else:
            print("未選擇檔案")  # 調試訊息

    # 綁定點擊事件，顯示右側結果
    def on_sheet_select(event):
        """當選擇工作表時，顯示右側結果"""
        global selected_file_global
        selected_sheet_index = sheet_listbox.curselection()
        
        if not selected_file_global:
            print("未選擇檔案")  # 調試訊息
            return
        
        if selected_sheet_index:
            selected_sheet = sheet_listbox.get(selected_sheet_index)
        else:
            print("未選擇工作表")  # 調試訊息
            return
        
        print(f"選擇的檔案: {selected_file_global}, 工作表: {selected_sheet}")  # 調試訊息
        if selected_file_global in grouped_results and selected_sheet in grouped_results[selected_file_global]:
            display_results(selected_file_global, selected_sheet, grouped_results)
        else:
            print(f"無對應結果: 檔案={selected_file_global}, 工作表={selected_sheet}")  # 調試訊息

    # 綁定事件
    file_listbox.bind("<<ListboxSelect>>", on_file_select)
    sheet_listbox.bind("<<ListboxSelect>>", on_sheet_select)

# 建立主視窗
root = tk.Tk()
root.title("Excel 關鍵字搜尋工具")

# 設定主視窗的行列權重，讓元件隨視窗縮放
root.grid_rowconfigure(4, weight=1)  # 第 4 行（結果顯示區域）可垂直縮放
root.grid_columnconfigure(0, weight=1)  # 第 0 列（左側文字框）可水平縮放
root.grid_columnconfigure(1, weight=1)  # 第 1 列（中間文字框）可水平縮放
root.grid_columnconfigure(2, weight=3)  # 第 2 列（右側表格）可水平縮放

# 資料夾路徑
folder_path_var = tk.StringVar()
tk.Label(root, text="選擇資料夾:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
tk.Entry(root, textvariable=folder_path_var, width=40).grid(row=0, column=1, padx=10, pady=5, sticky="we", columnspan=2)
tk.Button(root, text="瀏覽", command=browse_folder).grid(row=0, column=3, padx=10, pady=5)

# 第一個關鍵字
keyword1_var = tk.StringVar()
tk.Label(root, text="第一個關鍵字:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
tk.Entry(root, textvariable=keyword1_var, width=40).grid(row=1, column=1, padx=10, pady=5, sticky="we", columnspan=2)

# 第二個關鍵字
keyword2_var = tk.StringVar()
tk.Label(root, text="第二個關鍵字 (可選):").grid(row=2, column=0, padx=10, pady=5, sticky="e")
tk.Entry(root, textvariable=keyword2_var, width=40).grid(row=2, column=1, padx=10, pady=5, sticky="we", columnspan=2)

# 搜尋按鈕
tk.Button(root, text="開始搜尋", command=search_keywords).grid(row=3, column=0, columnspan=4, pady=10)

# 匯出按鈕
tk.Button(root, text="匯出搜尋結果", command=export_results).grid(row=3, column=2, columnspan=2, pady=10)

# 左側檔案名稱列表
file_listbox = tk.Listbox(root, width=30, height=20)
file_listbox.grid(row=4, column=0, padx=10, pady=10, sticky="nsew")

# 中間工作表名稱列表
sheet_listbox = tk.Listbox(root, width=30, height=20)
sheet_listbox.grid(row=4, column=1, padx=10, pady=10, sticky="nsew")

# 右側表格顯示結果
columns = ["箱號", "JAN", "數量", "品名", "客編", "客戶", "淨重", "毛重", "材積", "直派單號", "備註", "宅配單號", "宅配", "TIMES"]
tree = ttk.Treeview(root, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=100, anchor="center")
tree.grid(row=4, column=2, padx=10, pady=10, sticky="nsew", columnspan=2)

# 啟動主迴圈
root.mainloop()